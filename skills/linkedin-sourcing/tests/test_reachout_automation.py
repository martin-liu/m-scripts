#!/usr/bin/env python3
"""Tests for reachout_automation.py

Run with: python3 -m pytest skills/linkedin-sourcing/tests/test_reachout_automation.py -v
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

import reachout_automation as ra


class TestShouldExcludeByTitle:
    """Tests for title exclusion logic."""

    def test_exact_match(self):
        assert ra.should_exclude_by_title("Manager", "Manager") is True

    def test_case_insensitive_match(self):
        assert ra.should_exclude_by_title("SENIOR MANAGER", "manager") is True
        assert ra.should_exclude_by_title("senior manager", "Manager") is True

    def test_partial_match(self):
        assert (
            ra.should_exclude_by_title("Senior Engineering Manager", "Manager") is True
        )
        assert ra.should_exclude_by_title("Product Manager", "Manager,Director") is True

    def test_multiple_patterns(self):
        exclude = "Manager,Director,VP,Product Manager"
        assert ra.should_exclude_by_title("Senior Director", exclude) is True
        assert ra.should_exclude_by_title("VP of Engineering", exclude) is True
        assert ra.should_exclude_by_title("Product Manager", exclude) is True

    def test_no_match(self):
        assert ra.should_exclude_by_title("Software Engineer", "Manager") is False
        assert ra.should_exclude_by_title("Staff Engineer", "Manager,Director") is False

    def test_empty_title(self):
        assert ra.should_exclude_by_title(None, "Manager") is False
        assert ra.should_exclude_by_title("", "Manager") is False

    def test_empty_exclude_list(self):
        assert ra.should_exclude_by_title("Manager", "") is False


class TestExtractFirstName:
    """Tests for first name extraction."""

    def test_single_name(self):
        assert ra.extract_first_name("John") == "John"

    def test_full_name(self):
        assert ra.extract_first_name("John Smith") == "John"
        assert ra.extract_first_name("Mary Jane Watson") == "Mary"

    def test_empty_name(self):
        assert ra.extract_first_name(None) == "there"
        assert ra.extract_first_name("") == "there"


class TestSanitizeCoreFunction:
    """Tests for CORE_FUNCTION sanitization to ensure natural phrasing."""

    def test_removes_our_team_is_prefix(self):
        """Should remove 'Our team is' prefix."""
        result = ra.sanitize_core_function("Our team is building ML infrastructure")
        assert result == "building ML infrastructure"

    def test_removes_we_are_prefix(self):
        """Should remove 'We are' prefix."""
        result = ra.sanitize_core_function("We are dedicated to improving search")
        assert result == "dedicated to improving search"

    def test_removes_our_team_prefix(self):
        """Should remove 'Our team' prefix."""
        result = ra.sanitize_core_function("Our team builds scalable systems")
        assert result == "builds scalable systems"

    def test_removes_the_team_is_prefix(self):
        """Should remove 'The team is' prefix."""
        result = ra.sanitize_core_function("The team is creating AI solutions")
        assert result == "creating AI solutions"

    def test_lowercases_verb_phrases(self):
        """Should lowercase first letter for verb phrases after prepositions."""
        result = ra.sanitize_core_function("Building ML infrastructure")
        assert result == "building ML infrastructure"

    def test_preserves_proper_nouns(self):
        """Should preserve uppercase for likely proper nouns/acronyms."""
        result = ra.sanitize_core_function("Our team is building AI systems")
        assert "AI" in result  # Acronym should stay uppercase

    def test_handles_empty_input(self):
        """Should handle empty or None input gracefully."""
        assert ra.sanitize_core_function("") == ""
        assert ra.sanitize_core_function(None) == ""

    def test_handles_already_clean_input(self):
        """Should leave already-clean input unchanged."""
        result = ra.sanitize_core_function("improving search relevance")
        assert result == "improving search relevance"

    def test_simulation_exposed_case_building(self):
        """Test the exact awkward case from live simulation: 'complex Our team is building ... problems'."""
        raw_core_function = (
            "Our team is building scalable video codec hardware solutions"
        )
        sanitized = ra.sanitize_core_function(raw_core_function)
        # Should NOT contain "Our team is" prefix
        assert "Our team is" not in sanitized
        # Should read naturally in "complex ... problems" context
        full_sentence = f"tackling complex {sanitized} problems"
        assert "complex Our team" not in full_sentence
        assert "complex building" in full_sentence

    def test_simulation_exposed_case_dedicated_to(self):
        """Test the exact awkward case from live simulation: 'dedicated to Our team is building ...'."""
        raw_core_function = "Our team is building ML infrastructure from the ground up"
        sanitized = ra.sanitize_core_function(raw_core_function)
        # Should NOT contain "Our team is" prefix
        assert "Our team is" not in sanitized
        # Should read naturally after "dedicated to"
        full_sentence = f"We are dedicated to {sanitized}"
        assert "dedicated to Our team" not in full_sentence
        assert "dedicated to building" in full_sentence

    def test_second_pass_senior_personalization_reads_naturally(self):
        """Test that senior/staff/principal personalization reads naturally with sanitized CORE_FUNCTION.

        Issue: 'while tackling complex building ... problems' was awkward.
        Fix: Changed to 'while tackling the complex engineering challenges involved in ...'
        """
        config = {
            "TEAM_NAME": "AI Platform",
            "POSITION_TITLE": "ML Engineer",
            "CORE_FUNCTION": "Our team is building scalable ML infrastructure",
        }
        result = ra.generate_personalized_sentence(
            "Senior Engineer", "Company", "", None, config
        )
        # Should NOT have awkward "complex building ... problems"
        assert "complex building" not in result
        # Should have natural phrasing with "engineering challenges involved in"
        assert "complex engineering challenges involved in" in result
        # Should include the sanitized core function
        assert "building scalable ML infrastructure" in result


class TestGeneratePersonalizedSentence:
    """Tests for personalized sentence generation."""

    def test_pytorch_expertise(self):
        config = {
            "TEAM_NAME": "AI Platform",
            "POSITION_TITLE": "ML Engineer",
            "CORE_FUNCTION": "model training",
        }
        result = ra.generate_personalized_sentence(
            "Engineer", "Company", "PyTorch expert", None, config
        )
        assert "PyTorch" in result
        assert "AI Platform" in result

    def test_cuda_expertise(self):
        config = {
            "TEAM_NAME": "Infra",
            "POSITION_TITLE": "Engineer",
            "CORE_FUNCTION": "training",
        }
        result = ra.generate_personalized_sentence(
            "Engineer", "Company", "CUDA optimization", None, config
        )
        assert "CUDA" in result

    def test_top_company_background(self):
        config = {
            "TEAM_NAME": "AI",
            "POSITION_TITLE": "Engineer",
            "CORE_FUNCTION": "training",
        }
        result = ra.generate_personalized_sentence(
            "Engineer", "Google", "Software Engineer", None, config
        )
        assert "Google" in result
        assert "unique insights" in result

    def test_phd_background(self):
        config = {
            "TEAM_NAME": "Research",
            "POSITION_TITLE": "Scientist",
            "CORE_FUNCTION": "AI research",
        }
        result = ra.generate_personalized_sentence(
            "Researcher", "University", "PhD in ML", None, config
        )
        assert "research" in result.lower()

    def test_infrastructure_title(self):
        config = {
            "TEAM_NAME": "Platform",
            "POSITION_TITLE": "Engineer",
            "CORE_FUNCTION": "infrastructure",
        }
        result = ra.generate_personalized_sentence(
            "Infrastructure Engineer", "Company", "", None, config
        )
        assert "infrastructure" in result.lower()

    def test_senior_title(self):
        config = {
            "TEAM_NAME": "AI",
            "POSITION_TITLE": "Engineer",
            "CORE_FUNCTION": "training",
        }
        result = ra.generate_personalized_sentence(
            "Senior Engineer", "Company", "", None, config
        )
        assert "senior-level" in result.lower()

    def test_fallback_with_company_and_title(self):
        config = {
            "TEAM_NAME": "AI",
            "POSITION_TITLE": "Engineer",
            "CORE_FUNCTION": "training",
        }
        result = ra.generate_personalized_sentence(
            "Engineer", "Startup", "", None, config
        )
        assert "Startup" in result
        assert "Engineer" in result

    def test_fallback_with_headline(self):
        config = {
            "TEAM_NAME": "AI",
            "POSITION_TITLE": "Engineer",
            "CORE_FUNCTION": "training",
        }
        # Headline without ML keywords to test the headline fallback path
        result = ra.generate_personalized_sentence(
            "", "", "Building scalable data pipelines at scale", None, config
        )
        assert "scalable data pipelines" in result

    def test_ultimate_fallback(self):
        config = {
            "TEAM_NAME": "AI Platform",
            "POSITION_TITLE": "Engineer",
            "CORE_FUNCTION": "training",
        }
        result = ra.generate_personalized_sentence(None, None, None, None, config)
        assert "AI Platform" in result


class TestParseTemplate:
    """Tests for template parsing."""

    def test_parse_valid_template(self, tmp_path):
        template = tmp_path / "test_template.txt"
        template.write_text("""Subject: Hello {FirstName}

Hi {FirstName},

This is the body.

Best,
Team""")
        subject, body = ra.parse_template(str(template))
        assert subject == "Hello {FirstName}"
        assert "Hi {FirstName}," in body
        assert "This is the body." in body

    def test_parse_template_no_blank_line(self, tmp_path):
        template = tmp_path / "test_template.txt"
        template.write_text("""Subject: Test Subject
Body line 1
Body line 2""")
        subject, body = ra.parse_template(str(template))
        assert subject == "Test Subject"
        assert "Body line 1" in body

    def test_parse_missing_file(self, tmp_path):
        missing_file = tmp_path / "nonexistent.txt"
        try:
            ra.parse_template(str(missing_file))
            assert False, "Should have raised FileNotFoundError"
        except FileNotFoundError:
            pass


class TestFillTemplate:
    """Tests for template filling."""

    def test_fill_basic_placeholders(self):
        candidate = {
            "name": "John Smith",
            "title": "Senior Engineer",
            "company": "Google",
            "headline": "ML Infrastructure",
            "notes": "",
        }
        config = {
            "POSITION_TITLE": "ML Engineer",
            "TEAM_NAME": "AI Platform",
            "LOCATION": "San Francisco",
            "CORE_FUNCTION": "model training",
            "BUSINESS_IMPACT": "better AI products",
            "KEYWORDS": "PyTorch, TensorFlow",
            "USER_EMAIL": "hiring@company.com",
        }
        subject_template = "{FirstName}, {POSITION_TITLE} at {TEAM_NAME}"
        body_template = "Hi {FirstName}, your work at {Company} as {current_title} is impressive. {1 personalized sentence on why their background impressed you}"

        subject, body = ra.fill_template(
            subject_template, body_template, candidate, config
        )

        assert "John" in subject
        assert "ML Engineer" in subject
        assert "AI Platform" in subject
        assert "John" in body
        assert "Google" in body
        assert "Senior Engineer" in body

    def test_fill_preserves_existing_drafts(self):
        """Already drafted rows should not be re-drafted unless filter excludes them."""
        candidate = {
            "name": "Jane Doe",
            "title": "Engineer",
            "company": "Meta",
            "status": "Drafted",
            "next_action": "review",
            "draft_subject": "Existing Subject",
            "draft_body": "Existing Body",
        }
        config = {
            "POSITION_TITLE": "Engineer",
            "TEAM_NAME": "AI",
            "LOCATION": "SF",
            "CORE_FUNCTION": "AI",
            "BUSINESS_IMPACT": "impact",
            "KEYWORDS": "ML",
            "USER_EMAIL": "test@test.com",
        }

        subject_template = "Subject: New {FirstName}"
        body_template = "New body for {FirstName}"

        subject, body = ra.fill_template(
            subject_template, body_template, candidate, config
        )

        assert "Jane" in subject
        assert "Jane" in body

    def test_fill_template_sanitizes_core_function(self):
        """Template filling should use sanitized CORE_FUNCTION for natural phrasing."""
        candidate = {
            "name": "John Smith",
            "title": "Senior Engineer",
            "company": "Google",
            "headline": "",
            "notes": "",
        }
        config = {
            "POSITION_TITLE": "ML Engineer",
            "TEAM_NAME": "AI Platform",
            "LOCATION": "SF",
            "CORE_FUNCTION": "Our team is building scalable ML infrastructure",
            "BUSINESS_IMPACT": "improving user experience",
            "KEYWORDS": "PyTorch",
            "USER_EMAIL": "test@test.com",
        }

        # Template that would create awkward phrasing with raw CORE_FUNCTION
        subject_template = "Subject: {FirstName}, {POSITION_TITLE}"
        body_template = "We are dedicated to {CORE_FUNCTION}. Join us!"

        subject, body = ra.fill_template(
            subject_template, body_template, candidate, config
        )

        # Should NOT have awkward "dedicated to Our team is building"
        assert "dedicated to Our team is" not in body
        # Should have natural "dedicated to building"
        assert "dedicated to building" in body

    def test_fill_template_core_function_in_personalized_sentence(self):
        """Personalized sentences should use sanitized CORE_FUNCTION."""
        candidate = {
            "name": "Jane Doe",
            "title": "Staff Engineer",
            "company": "Acme Silicon",
            "headline": "",
            "notes": "",
        }
        config = {
            "POSITION_TITLE": "Engineer",
            "TEAM_NAME": "AI",
            "LOCATION": "SF",
            "CORE_FUNCTION": "Our team is building distributed training systems",
            "BUSINESS_IMPACT": "impact",
            "KEYWORDS": "ML",
            "USER_EMAIL": "test@test.com",
        }

        # Template with personalized sentence that uses CORE_FUNCTION
        subject_template = "Subject: {FirstName}"
        body_template = (
            "{1 personalized sentence on why their background impressed you}"
        )

        subject, body = ra.fill_template(
            subject_template, body_template, candidate, config
        )

        # Should NOT have awkward "complex Our team is building ... problems"
        assert "complex Our team is" not in body
        # Should have natural phrasing with "engineering challenges involved in"
        assert "complex engineering challenges involved in" in body

    def test_second_pass_business_impact_verb_phrase(self):
        """Test that BUSINESS_IMPACT works with verb phrases like 'better serve billions of users'.

        Issue: 'delivering better serve billions of users' was grammatically incorrect.
        Fix: Changed template to 'we aim to {BUSINESS_IMPACT}'.
        """
        candidate = {
            "name": "John Smith",
            "title": "Senior Engineer",
            "company": "Google",
            "headline": "",
            "notes": "",
        }
        config = {
            "POSITION_TITLE": "ML Engineer",
            "TEAM_NAME": "AI Platform",
            "LOCATION": "San Francisco",
            "CORE_FUNCTION": "building scalable ML infrastructure",
            "BUSINESS_IMPACT": "better serve billions of users",
            "KEYWORDS": "PyTorch",
            "USER_EMAIL": "test@test.com",
        }

        # Template using the new "we aim to" phrasing
        subject_template = "Subject: {FirstName}, {POSITION_TITLE}"
        body_template = "Our team is dedicated to {CORE_FUNCTION}, and we aim to {BUSINESS_IMPACT} for our users."

        subject, body = ra.fill_template(
            subject_template, body_template, candidate, config
        )

        # Should NOT have awkward "delivering better serve"
        assert "delivering better serve" not in body
        # Should have natural "we aim to better serve"
        assert "we aim to better serve billions of users" in body
        # Full sentence should read naturally
        assert "dedicated to building scalable ML infrastructure, and we aim to" in body


class TestParseConfig:
    """Tests for config file parsing."""

    def test_parse_valid_config(self, tmp_path):
        config_file = tmp_path / "config.sh"
        config_file.write_text("""# Project config
PROJECT_ID="123"
POSITION_TITLE="Engineer"
TEAM_NAME="AI Team"
EXCLUDE_TITLES="Manager,Director"
DAILY_LIMIT=200
""")
        config = ra.parse_config(str(config_file))
        assert config["PROJECT_ID"] == "123"
        assert config["POSITION_TITLE"] == "Engineer"
        assert config["TEAM_NAME"] == "AI Team"
        assert config["EXCLUDE_TITLES"] == "Manager,Director"
        assert config["DAILY_LIMIT"] == "200"

    def test_parse_config_with_spaces(self, tmp_path):
        config_file = tmp_path / "config.sh"
        config_file.write_text(
            'POSITION_TITLE="Senior Engineer"\nLOCATION="San Francisco, CA"'
        )
        config = ra.parse_config(str(config_file))
        assert config["POSITION_TITLE"] == "Senior Engineer"
        assert config["LOCATION"] == "San Francisco, CA"

    def test_parse_missing_file(self, tmp_path):
        missing_file = tmp_path / "nonexistent.sh"
        try:
            ra.parse_config(str(missing_file))
            assert False, "Should have raised FileNotFoundError"
        except FileNotFoundError:
            pass


class TestLoadMergedConfig:
    """Tests for load_merged_config function."""

    def test_project_config_only(self, tmp_path, monkeypatch):
        """When no profile config exists, use project config only."""
        monkeypatch.setattr(ra.Path, "home", lambda: tmp_path)  # Mock home to tmp_path

        config_file = tmp_path / "project.sh"
        config_file.write_text('POSITION_TITLE="Engineer"\nTEAM_NAME="AI Team"')

        config = ra.load_merged_config(str(config_file))
        assert config["POSITION_TITLE"] == "Engineer"
        assert config["TEAM_NAME"] == "AI Team"

    def test_profile_config_fallback(self, tmp_path, monkeypatch):
        """Profile config provides fallback for missing keys."""
        # Create profile config directory
        profile_dir = tmp_path / ".config" / "linkedin-sourcing"
        profile_dir.mkdir(parents=True)
        profile_file = profile_dir / "profile.sh"
        profile_file.write_text(
            'USER_EMAIL="profile@example.com"\nTEAM_NAME="Profile Team"'
        )

        # Create project config
        project_file = tmp_path / "project.sh"
        project_file.write_text('POSITION_TITLE="Engineer"')

        monkeypatch.setattr(ra.Path, "home", lambda: tmp_path)

        config = ra.load_merged_config(str(project_file))
        assert config["POSITION_TITLE"] == "Engineer"  # From project
        assert config["USER_EMAIL"] == "profile@example.com"  # From profile
        assert config["TEAM_NAME"] == "Profile Team"  # From profile

    def test_project_config_takes_precedence(self, tmp_path, monkeypatch):
        """Project config values override profile config."""
        # Create profile config
        profile_dir = tmp_path / ".config" / "linkedin-sourcing"
        profile_dir.mkdir(parents=True)
        profile_file = profile_dir / "profile.sh"
        profile_file.write_text(
            'TEAM_NAME="Profile Team"\nUSER_EMAIL="profile@example.com"'
        )

        # Create project config with same key
        project_file = tmp_path / "project.sh"
        project_file.write_text('TEAM_NAME="Project Team"')

        monkeypatch.setattr(ra.Path, "home", lambda: tmp_path)

        config = ra.load_merged_config(str(project_file))
        assert config["TEAM_NAME"] == "Project Team"  # Project wins
        assert config["USER_EMAIL"] == "profile@example.com"  # From profile

    def test_user_email_from_profile(self, tmp_path, monkeypatch):
        """USER_EMAIL can come from profile config when not in project config."""
        profile_dir = tmp_path / ".config" / "linkedin-sourcing"
        profile_dir.mkdir(parents=True)
        profile_file = profile_dir / "profile.sh"
        profile_file.write_text('USER_EMAIL="recruiter@company.com"')

        project_file = tmp_path / "project.sh"
        project_file.write_text('POSITION_TITLE="Engineer"')

        monkeypatch.setattr(ra.Path, "home", lambda: tmp_path)

        config = ra.load_merged_config(str(project_file))
        assert config["USER_EMAIL"] == "recruiter@company.com"

        # Verify it works in template filling
        candidate = {
            "name": "John",
            "title": "Engineer",
            "company": "Google",
            "headline": "",
            "notes": "",
        }
        subject_template = "Hello"
        body_template = "Contact me at {USER_EMAIL}"
        subject, body = ra.fill_template(
            subject_template, body_template, candidate, config
        )
        assert "recruiter@company.com" in body


class TestSchemaMigration:
    """Tests for workbook schema migration and backward compatibility."""

    def create_old_schema_workbook(self, tmp_path, rows_data):
        """Create a workbook with old 18-column schema (no headline/location)."""
        Workbook = ra.get_openpyxl().Workbook

        old_columns = [
            "row_id",
            "name",
            "company",
            "title",
            "profile_url",
            "est_yoe",
            "highest_degree",
            "school",
            "status",
            "next_action",
            "draft_subject",
            "draft_body",
            "date_sent",
            "attempts",
            "last_contact",
            "reply_type",
            "reply_summary",
            "notes",
        ]

        wb_path = tmp_path / "test_old.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        ws.append(old_columns)

        for i, row_data in enumerate(rows_data, start=1):
            row = [i] + [row_data.get(col, "") for col in old_columns[1:]]
            ws.append(row)

        wb.save(wb_path)
        return str(wb_path)

    def test_read_old_workbook_adds_missing_columns(self, tmp_path):
        """Reading old workbooks should add missing columns as None."""
        rows = [
            {
                "name": "John",
                "title": "Engineer",
                "company": "Google",
                "status": "Extracted",
                "next_action": "draft",
                "notes": "PyTorch expert with CUDA experience",
            },
        ]

        wb_path = self.create_old_schema_workbook(tmp_path, rows)

        # Load without migration to test read resilience
        wb = ra.load_workbook(wb_path, migrate_schema=False)
        read_rows = ra.read_all_rows(wb)

        # Should have headline and location as None
        assert len(read_rows) == 1
        assert read_rows[0]["headline"] is None
        assert read_rows[0]["location"] is None
        assert read_rows[0]["name"] == "John"
        assert read_rows[0]["notes"] == "PyTorch expert with CUDA experience"

    def test_ensure_schema_adds_missing_columns(self, tmp_path):
        """ensure_schema should add headline and location to old workbooks."""
        from excel_utils import ensure_schema, COLUMNS

        rows = [{"name": "Jane", "title": "Engineer", "status": "Extracted"}]
        wb_path = self.create_old_schema_workbook(tmp_path, rows)

        # Run migration
        migrated = ensure_schema(wb_path)
        assert migrated is True

        # Load and verify new columns exist
        wb = ra.load_workbook(wb_path, migrate_schema=False)
        ws = wb["Candidates"]
        headers = [cell.value for cell in ws[1] if cell.value]

        assert "headline" in headers
        assert "location" in headers
        assert len(headers) == len(COLUMNS)

    def test_ensure_schema_idempotent(self, tmp_path):
        """ensure_schema should return False if no migration needed."""
        from excel_utils import ensure_schema

        # Create new schema workbook
        rows = [{"name": "Bob", "title": "Engineer", "status": "Extracted"}]
        wb_path = tmp_path / "test_new.xlsx"
        wb = ra.get_openpyxl().Workbook()
        ws = wb.active
        ws.title = "Candidates"
        ws.append(ra.COLUMNS)
        for i, row_data in enumerate(rows, start=1):
            row = [i] + [row_data.get(col, "") for col in ra.COLUMNS[1:]]
            ws.append(row)
        wb.save(wb_path)

        # Second migration should return False
        migrated = ensure_schema(wb_path)
        assert migrated is False


class TestFallbackPersonalization:
    """Tests for personalization fallback when headline is missing."""

    def test_fallback_to_notes_when_headline_empty(self):
        """Personalization should use notes when headline is empty."""
        config = {
            "TEAM_NAME": "AI Platform",
            "POSITION_TITLE": "ML Engineer",
            "CORE_FUNCTION": "model training",
        }
        # Empty headline, notes contains PyTorch
        result = ra.generate_personalized_sentence(
            "Engineer",
            "Company",
            "",
            "PyTorch expert with distributed training",
            config,
        )
        assert "PyTorch" in result

    def test_fallback_to_notes_when_headline_none(self):
        """Personalization should use notes when headline is None (old workbook)."""
        config = {
            "TEAM_NAME": "AI Platform",
            "POSITION_TITLE": "ML Engineer",
            "CORE_FUNCTION": "model training",
        }
        result = ra.generate_personalized_sentence(
            "Engineer", "Company", None, "CUDA optimization expert", config
        )
        assert "CUDA" in result

    def test_fill_template_uses_notes_fallback(self):
        """fill_template should fallback to notes for personalization."""
        candidate = {
            "name": "John Smith",
            "title": "Engineer",
            "company": "Google",
            "headline": "",  # Empty - should fallback to notes
            "notes": "PyTorch and CUDA expert",
        }
        config = {
            "POSITION_TITLE": "ML Engineer",
            "TEAM_NAME": "AI Platform",
            "LOCATION": "SF",
            "CORE_FUNCTION": "training",
            "BUSINESS_IMPACT": "impact",
            "KEYWORDS": "ML",
            "USER_EMAIL": "test@test.com",
        }

        subject_template = "Hello {FirstName}"
        body_template = (
            "{1 personalized sentence on why their background impressed you}"
        )

        subject, body = ra.fill_template(
            subject_template, body_template, candidate, config
        )

        # Should mention PyTorch from notes since headline is empty
        assert "PyTorch" in body


class TestIntegrationWithExcel:
    """Integration tests using actual Excel files."""

    def create_test_workbook(self, tmp_path, rows_data):
        """Helper to create a test workbook with candidate data."""
        from openpyxl import Workbook

        wb_path = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        ws.append(ra.COLUMNS)

        for i, row_data in enumerate(rows_data, start=1):
            row = [i] + [row_data.get(col, "") for col in ra.COLUMNS[1:]]
            ws.append(row)

        wb.save(wb_path)
        return str(wb_path)

    def test_filter_command(self, tmp_path):
        """Test the filter command excludes correct titles."""
        config_file = tmp_path / "config.sh"
        config_file.write_text(
            'EXCLUDE_TITLES="Manager,Director,VP"\nPOSITION_TITLE="Engineer"\nTEAM_NAME="AI"\nLOCATION="SF"\nCORE_FUNCTION="training"\nBUSINESS_IMPACT="impact"\nUSER_EMAIL="test@test.com"'
        )

        rows = [
            {
                "name": "John",
                "title": "Software Engineer",
                "company": "Google",
                "status": "Extracted",
                "next_action": "filter",
            },
            {
                "name": "Jane",
                "title": "Engineering Manager",
                "company": "Meta",
                "status": "Extracted",
                "next_action": "filter",
            },
            {
                "name": "Bob",
                "title": "Senior Director",
                "company": "Apple",
                "status": "Extracted",
                "next_action": "draft",
            },
            {
                "name": "Alice",
                "title": "VP Product",
                "company": "Netflix",
                "status": "Extracted",
                "next_action": "",
            },
        ]

        wb_path = self.create_test_workbook(tmp_path, rows)
        result = ra.cmd_filter(wb_path, str(config_file))

        assert result["kept"] == 1
        assert result["filtered"] == 3

        wb = ra.load_workbook(wb_path)
        updated_rows = ra.read_all_rows(wb)

        john_row = next(r for r in updated_rows if r["name"] == "John")
        assert john_row["next_action"] == "draft"

        jane_row = next(r for r in updated_rows if r["name"] == "Jane")
        assert jane_row["status"] == "Filtered"
        assert jane_row["next_action"] == "done"

    def test_filter_preserves_already_drafted(self):
        """Already drafted rows should not be modified by filter."""
        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            config_file = tmp_path / "config.sh"
            config_file.write_text(
                'EXCLUDE_TITLES="Manager"\nPOSITION_TITLE="Engineer"\nTEAM_NAME="AI"\nLOCATION="SF"\nCORE_FUNCTION="training"\nBUSINESS_IMPACT="impact"\nUSER_EMAIL="test@test.com"'
            )

            rows = [
                {
                    "name": "John",
                    "title": "Engineer",
                    "status": "Drafted",
                    "next_action": "review",
                    "draft_subject": "Existing",
                },
                {
                    "name": "Jane",
                    "title": "Manager",
                    "status": "Extracted",
                    "next_action": "filter",
                },
            ]

            wb_path = self.create_test_workbook(tmp_path, rows)
            result = ra.cmd_filter(wb_path, str(config_file))

            wb = ra.load_workbook(wb_path)
            updated_rows = ra.read_all_rows(wb)

            john_row = next(r for r in updated_rows if r["name"] == "John")
            assert john_row["status"] == "Drafted"
            assert john_row["draft_subject"] == "Existing"

    def test_draft_command(self, tmp_path):
        """Test the draft command generates personalized content."""
        template_file = tmp_path / "template.txt"
        template_file.write_text("""Subject: {FirstName}, {POSITION_TITLE} Opportunity

Hi {FirstName},

Your role as {current_title} at {Company} caught my eye. {1 personalized sentence on why their background impressed you}

Best,
Team""")

        config_file = tmp_path / "config.sh"
        config_file.write_text(
            'POSITION_TITLE="ML Engineer"\nTEAM_NAME="AI Platform"\nLOCATION="SF"\nCORE_FUNCTION="training"\nBUSINESS_IMPACT="impact"\nUSER_EMAIL="test@test.com"'
        )

        rows = [
            {
                "name": "John Smith",
                "title": "PyTorch Engineer",
                "company": "OpenAI",
                "status": "Extracted",
                "next_action": "draft",
                "headline": "PyTorch expert",
                "notes": "",
            },
        ]

        wb_path = self.create_test_workbook(tmp_path, rows)
        result = ra.cmd_draft(wb_path, str(config_file), str(template_file))

        assert result["drafted"] == 1

        wb = ra.load_workbook(wb_path)
        updated_rows = ra.read_all_rows(wb)

        john_row = updated_rows[0]
        assert john_row["status"] == "Drafted"
        assert john_row["next_action"] == "review"
        assert "John" in john_row["draft_subject"]
        assert "PyTorch" in john_row["draft_body"]

    def test_draft_skips_already_drafted(self, tmp_path):
        """Draft command should skip rows already in Drafted status."""
        template_file = tmp_path / "template.txt"
        template_file.write_text("Subject: Test\n\nBody")

        config_file = tmp_path / "config.sh"
        config_file.write_text(
            'POSITION_TITLE="Engineer"\nTEAM_NAME="AI"\nLOCATION="SF"\nCORE_FUNCTION="training"\nBUSINESS_IMPACT="impact"\nUSER_EMAIL="test@test.com"'
        )

        rows = [
            {
                "name": "John",
                "title": "Engineer",
                "status": "Drafted",
                "next_action": "review",
                "draft_subject": "Existing",
            },
            {
                "name": "Jane",
                "title": "Engineer",
                "status": "Extracted",
                "next_action": "draft",
            },
        ]

        wb_path = self.create_test_workbook(tmp_path, rows)
        result = ra.cmd_draft(wb_path, str(config_file), str(template_file))

        assert result["drafted"] == 1
        assert result["skipped"] == 1

    def test_approve_command(self, tmp_path):
        """Test the approve command updates drafted rows."""
        rows = [
            {
                "name": "John",
                "title": "Engineer",
                "status": "Drafted",
                "next_action": "review",
            },
            {
                "name": "Jane",
                "title": "Engineer",
                "status": "Extracted",
                "next_action": "draft",
            },
            {
                "name": "Bob",
                "title": "Manager",
                "status": "Filtered",
                "next_action": "done",
            },
        ]

        wb_path = self.create_test_workbook(tmp_path, rows)
        result = ra.cmd_approve(wb_path)

        assert result["approved"] == 1
        assert result["skipped"] == 2

        wb = ra.load_workbook(wb_path)
        updated_rows = ra.read_all_rows(wb)

        john_row = next(r for r in updated_rows if r["name"] == "John")
        assert john_row["status"] == "Approved"
        assert john_row["next_action"] == "send"

    def test_summary_command(self, tmp_path):
        """Test the summary command returns correct counts."""
        rows = [
            {"name": "John", "status": "Extracted", "next_action": "filter"},
            {"name": "Jane", "status": "Drafted", "next_action": "review"},
            {"name": "Bob", "status": "Approved", "next_action": "send"},
            {"name": "Alice", "status": "Sent", "next_action": "done"},
        ]

        wb_path = self.create_test_workbook(tmp_path, rows)
        result = ra.cmd_summary(wb_path)

        assert result["total"] == 4
        assert result["status_counts"]["Extracted"] == 1
        assert result["status_counts"]["Drafted"] == 1
        assert result["status_counts"]["Approved"] == 1
        assert result["status_counts"]["Sent"] == 1


if __name__ == "__main__":
    import pytest

    pytest.main([__file__, "-v"])
