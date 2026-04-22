#!/usr/bin/env python3
"""Reachout automation for LinkedIn sourcing - handles filter, draft, approve, summary phases.

Part of the workflow: Extract -> Filter -> Enrich -> Draft -> Review -> Send

Usage:
    python3 reachout_automation.py filter <workbook_path> <config_path> [--no-enrichment]
    python3 reachout_automation.py draft <workbook_path> <config_path> <template_path>
    python3 reachout_automation.py approve <workbook_path>
    python3 reachout_automation.py summary <workbook_path>

Commands:
    filter   - Exclude candidates by title, route kept rows to enrich (or draft with --no-enrichment)
    draft    - Generate personalized inmail drafts from templates (uses enrichment_notes if present)
    approve  - Auto-approve all Drafted rows for testing
    summary  - Print counts by status and next_action

Options:
    --no-enrichment  - Route filtered-in rows directly to draft (legacy behavior)
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

from config_utils import get_unresolved_project_messaging_fields
from typing import Any

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent

# Import COLUMNS from excel_utils for single source of truth
sys.path.insert(0, str(SCRIPT_DIR))
from excel_utils import COLUMNS, ensure_schema
from config_utils import parse_config_file


def get_openpyxl():
    """Lazy-load openpyxl, auto-installing if missing."""
    try:
        import openpyxl

        return openpyxl
    except ImportError:
        import subprocess

        try:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
            )
        except subprocess.CalledProcessError:
            subprocess.check_call(
                [
                    sys.executable,
                    "-m",
                    "pip",
                    "install",
                    "openpyxl",
                    "-q",
                    "--break-system-packages",
                ],
            )
        import openpyxl

        return openpyxl


def load_workbook(path: str, migrate_schema: bool = True):
    """Load an Excel workbook, optionally migrating schema if needed."""
    openpyxl = get_openpyxl()
    # Ensure schema is up to date before loading (for old workbooks)
    if migrate_schema:
        ensure_schema(path)
    return openpyxl.load_workbook(path)


def save_workbook(wb, path: str):
    """Save an Excel workbook."""
    wb.save(path)


def read_all_rows(wb) -> list[dict[str, Any]]:
    """Read all candidate rows from the workbook.

    Handles both old (18-column) and new (20-column) schemas gracefully.
    Missing columns return None instead of causing errors.
    """
    ws = wb["Candidates"]

    # Read actual headers from the workbook
    actual_headers = []
    for cell in ws[1]:
        if cell.value is not None:
            actual_headers.append(cell.value)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is None:
            continue
        # Map by actual header position
        data = {}
        for i, header in enumerate(actual_headers):
            data[header] = row[i].value if i < len(row) else None
        # Ensure all canonical columns exist (as None if missing from old workbook)
        for col in COLUMNS:
            if col not in data:
                data[col] = None
        rows.append(data)
    return rows


def update_row(wb, row_id: int, updates: dict[str, Any]):
    """Update a single row by row_id with the given column values."""
    ws = wb["Candidates"]
    actual_headers = []
    for cell in ws[1]:
        if cell.value is not None:
            actual_headers.append(cell.value)
    header_to_index = {header: idx for idx, header in enumerate(actual_headers)}
    for row in ws.iter_rows(min_row=2):
        if row[0].value == row_id:
            for col_name, value in updates.items():
                col_idx = header_to_index.get(col_name)
                if col_idx is not None:
                    row[col_idx].value = value
            return True
    return False


def parse_config(config_path: str) -> dict[str, str]:
    """Parse a bash-style config file into a dictionary.

    This is a thin wrapper around config_utils.parse_config_file that raises
    FileNotFoundError for missing files (for backward compatibility).
    """
    path = Path(config_path)
    if not path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")
    return parse_config_file(config_path)


def load_merged_config(project_config_path: str) -> dict[str, str]:
    """Load project config merged with profile config fallback.

    Profile config (at ~/.config/linkedin-sourcing/profile.sh) provides
    user-specific settings like USER_EMAIL that may not be in project config.
    Project config values take precedence over profile config.
    """
    # Start with project config
    config = parse_config(project_config_path)

    # Load profile config as fallback
    profile_path = Path.home() / ".config" / "linkedin-sourcing" / "profile.sh"
    if profile_path.exists():
        profile_config = parse_config(str(profile_path))
        # Merge: project config takes precedence, profile provides defaults
        for key, value in profile_config.items():
            if key not in config:
                config[key] = value

    return config


def get_matching_exclude_title_patterns(
    title: str | None,
    exclude_titles: str,
) -> list[str]:
    """Return the configured exclusion patterns that match a title."""
    if not title:
        return []

    title_lower = title.lower()
    matches: list[str] = []
    for raw_pattern in exclude_titles.split(","):
        pattern = raw_pattern.strip()
        if pattern and pattern.lower() in title_lower:
            matches.append(pattern)
    return matches


def should_exclude_by_title(title: str | None, exclude_titles: str) -> bool:
    """Check if a title should be excluded based on EXCLUDE_TITLES."""
    return bool(get_matching_exclude_title_patterns(title, exclude_titles))


def extract_first_name(name: str | None) -> str:
    """Extract first name from full name."""
    if not name:
        return "there"
    return name.split()[0]


def sanitize_core_function(core_function: str | None) -> str:
    """Sanitize CORE_FUNCTION for natural insertion into templates.

    Removes leading phrases like "Our team is building..." to convert them
    into forms that read naturally after "dedicated to" or in "complex ... problems".

    Examples:
        "Our team is building ML infrastructure" -> "building ML infrastructure"
        "We are dedicated to improving search" -> "improving search"
        "Building scalable systems" -> "building scalable systems"
    """
    if not core_function:
        return ""

    # Patterns that create awkward phrasing when inserted into templates
    leading_patterns = [
        r"^[Oo]ur team is\s+",
        r"^[Ww]e are\s+",
        r"^[Oo]ur team\s+",
        r"^[Tt]he team is\s+",
    ]

    result = core_function.strip()
    for pattern in leading_patterns:
        result = re.sub(pattern, "", result)

    # Lowercase first letter if it starts a verb phrase (for "dedicated to" context)
    # But preserve if it's a proper noun or starts with uppercase acronym
    if result and result[0].isupper() and not result[:2].isupper():
        # Check if it looks like a verb phrase that should be lowercased
        # after "dedicated to" or similar prepositions
        first_word = result.split()[0] if result.split() else ""
        if first_word.lower() in [
            "building",
            "creating",
            "developing",
            "improving",
            "designing",
            "implementing",
            "scaling",
            "optimizing",
        ]:
            result = result[0].lower() + result[1:]

    return result.strip()


def generate_personalized_sentence(
    title: str | None,
    company: str | None,
    headline: str | None,
    notes: str | None,
    config: dict[str, str],
) -> str:
    """Generate a genuinely specific personalized sentence based on candidate data.

    Uses actual candidate data (title, company, headline, notes) to create
    meaningful personalization, not generic filler.
    """
    title = title or ""
    company = company or ""
    headline = headline or ""
    notes = notes or ""

    team_name = config.get("TEAM_NAME", "")
    position_title = config.get("POSITION_TITLE", "")
    core_function = sanitize_core_function(config.get("CORE_FUNCTION", ""))

    patterns = []

    if "pytorch" in headline.lower() or "pytorch" in notes.lower():
        patterns.append(
            f"Your deep expertise with PyTorch and distributed training frameworks is exactly what our {team_name} team needs right now"
        )

    if (
        "cuda" in headline.lower()
        or "cuda" in notes.lower()
        or "gpu" in headline.lower()
    ):
        patterns.append(
            f"Your hands-on experience with CUDA and GPU optimization aligns perfectly with our infrastructure challenges"
        )

    if (
        "google" in company.lower()
        or "meta" in company.lower()
        or "openai" in company.lower()
    ):
        patterns.append(
            f"Your background at {company} gives you unique insights into large-scale AI systems that would be invaluable to our team"
        )

    if (
        "phd" in headline.lower()
        or "phd" in notes.lower()
        or "research" in headline.lower()
    ):
        patterns.append(
            f"Your research background and technical depth in {core_function} would make you a strong contributor to our most challenging projects"
        )

    if "infrastructure" in title.lower() or "platform" in title.lower():
        patterns.append(
            f"Your experience building {title.lower()} solutions directly maps to the scalability challenges we're solving"
        )

    if (
        "senior" in title.lower()
        or "staff" in title.lower()
        or "principal" in title.lower()
    ):
        patterns.append(
            f"Your senior-level experience would help mentor our growing team while tackling the complex engineering challenges involved in {core_function}"
        )

    if "machine learning" in headline.lower() or "ml" in headline.lower():
        patterns.append(
            f"Your ML systems background combined with our {position_title} focus creates a compelling match"
        )

    if patterns:
        return patterns[0]

    if company and title:
        return f"Your work as {title} at {company} demonstrates the kind of hands-on technical expertise we're looking for in this role"

    if headline:
        headline_clean = headline.split("·")[0].strip()
        if len(headline_clean) > 10:
            return f"Your background in {headline_clean} shows you have the relevant experience for this {position_title} opportunity"

    return f"Your profile suggests strong alignment with the technical challenges we're tackling in {team_name}"


def parse_template(template_path: str) -> tuple[str, str]:
    """Parse an inmail template file into subject and body.

    Line 1 should be: Subject: ...
    Blank line follows, then body content.
    """
    path = Path(template_path)
    if not path.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")

    content = path.read_text()
    lines = content.split("\n")

    subject = ""
    body_start = 0

    for i, line in enumerate(lines):
        if line.startswith("Subject:"):
            subject = line.replace("Subject:", "").strip()
            body_start = i + 1
            break

    while body_start < len(lines) and lines[body_start].strip() == "":
        body_start += 1

    body = "\n".join(lines[body_start:])
    return subject, body


def fill_template(
    template_subject: str,
    template_body: str,
    candidate: dict[str, Any],
    config: dict[str, str],
) -> tuple[str, str]:
    """Fill template placeholders with candidate data and config values.

    Uses headline for personalization when available, falls back to notes
    for old workbooks where headline column may be missing/empty.
    """
    first_name = extract_first_name(candidate.get("name"))
    title = candidate.get("title") or ""
    company = candidate.get("company") or ""
    # Fallback: use notes if headline is missing (old workbooks) or empty
    headline = candidate.get("headline") or ""
    notes = candidate.get("notes") or ""

    # Build effective headline for personalization: prefer headline, fallback to notes
    effective_headline = headline if headline else notes

    personalized = generate_personalized_sentence(
        title, company, effective_headline, notes, config
    )

    relevant_skills = config.get("KEYWORDS", "relevant technologies")
    if "pytorch" in effective_headline.lower() or "pytorch" in notes.lower():
        relevant_skills = "PyTorch, distributed training, and performance optimization"
    elif "cuda" in effective_headline.lower() or "gpu" in effective_headline.lower():
        relevant_skills = "CUDA, GPU kernels, and low-level optimization"
    elif "infrastructure" in title.lower():
        relevant_skills = "ML infrastructure and distributed systems"

    placeholders = {
        "{FirstName}": first_name,
        "{current_title}": title,
        "{Company}": company,
        "{POSITION_TITLE}": config.get("POSITION_TITLE", ""),
        "{TEAM_NAME}": config.get("TEAM_NAME", ""),
        "{LOCATION}": config.get("LOCATION", ""),
        "{CORE_FUNCTION}": sanitize_core_function(config.get("CORE_FUNCTION", "")),
        "{BUSINESS_IMPACT}": config.get("BUSINESS_IMPACT", ""),
        "{relevant_skills}": relevant_skills,
        "{USER_EMAIL}": config.get("USER_EMAIL", ""),
        "{1 personalized sentence on why their background impressed you}": personalized,
    }

    subject = template_subject
    body = template_body

    for placeholder, value in placeholders.items():
        subject = subject.replace(placeholder, str(value))
        body = body.replace(placeholder, str(value))

    return subject, body


def cmd_filter(workbook_path: str, config_path: str, use_enrichment: bool = True):
    """Filter candidates by title exclusion.

    Processes rows where next_action indicates filtering is needed
    (either 'filter' or when status=Extracted but next_action=draft).
    Excluded rows get status=Filtered, next_action=done.
    Kept rows get next_action=enrich (if use_enrichment) or next_action=draft.

    Args:
        workbook_path: Path to candidate workbook
        config_path: Path to project config
        use_enrichment: If True, route kept rows to enrichment phase.
                       If False, route directly to draft (legacy behavior).
    """
    config = load_merged_config(config_path)
    exclude_titles = config.get("EXCLUDE_TITLES", "")

    wb = load_workbook(workbook_path)
    rows = read_all_rows(wb)

    kept = 0
    filtered = 0
    skipped = 0
    filtered_details = []
    exclusion_patterns = [
        pattern.strip() for pattern in exclude_titles.split(",") if pattern.strip()
    ]

    for row in rows:
        status = row.get("status") or ""
        next_action = row.get("next_action") or ""
        row_id = row.get("row_id")

        if next_action == "filter":
            needs_filter = True
        elif status == "Extracted" and next_action == "draft":
            needs_filter = True
        elif status == "Extracted" and next_action == "enrich":
            # Already routed to enrichment - skip re-processing
            skipped += 1
            continue
        elif status == "Extracted" and not next_action:
            needs_filter = True
        else:
            needs_filter = False

        if not needs_filter:
            skipped += 1
            continue

        title = row.get("title") or ""

        matched_patterns = get_matching_exclude_title_patterns(title, exclude_titles)
        if matched_patterns:
            update_row(
                wb,
                row_id,
                {
                    "status": "Filtered",
                    "next_action": "done",
                },
            )
            filtered += 1
            filtered_details.append(
                {
                    "row_id": row_id,
                    "name": row.get("name") or "",
                    "title": title,
                    "matched_exclusion_rules": matched_patterns,
                    "reason": f"Matched exclusion rule(s): {', '.join(matched_patterns)}",
                }
            )
        else:
            # Route to enrichment or draft based on use_enrichment flag
            next_action_value = "enrich" if use_enrichment else "draft"
            update_row(
                wb,
                row_id,
                {
                    "next_action": next_action_value,
                },
            )
            kept += 1

    save_workbook(wb, workbook_path)

    target_phase = "enrich" if use_enrichment else "draft"
    print(
        f"Filter complete: {kept} kept (→{target_phase}), {filtered} filtered, {skipped} skipped"
    )
    return {
        "kept": kept,
        "filtered": filtered,
        "skipped": skipped,
        "target_phase": target_phase,
        "exclude_titles": exclusion_patterns,
        "filtered_details": filtered_details,
    }


def cmd_draft(
    workbook_path: str,
    config_path: str,
    template_path: str,
    row_ids: list[int] | None = None,
):
    """Generate draft inmails for candidates ready for drafting.

    Processes rows where next_action=draft or (next_action=enrich and already enriched).
    Generates personalized subject/body from template.
    Uses enrichment_notes when available for better personalization.
    Updates status=Drafted, next_action=review.
    """
    config = load_merged_config(config_path)

    placeholder_fields = get_unresolved_project_messaging_fields(config)
    if placeholder_fields:
        return {
            "drafted": 0,
            "skipped": 0,
            "error": (
                "Config contains placeholder values that must be updated before drafting: "
                f"{', '.join(placeholder_fields)}"
            ),
        }

    template_subject, template_body = parse_template(template_path)

    wb = load_workbook(workbook_path)
    rows = read_all_rows(wb)

    drafted = 0
    skipped = 0

    for row in rows:
        next_action = row.get("next_action") or ""
        status = row.get("status") or ""
        row_id = row.get("row_id")

        if row_ids and row_id not in row_ids:
            skipped += 1
            continue

        # Allow drafting for:
        # 1. next_action=draft (standard path)
        # 2. next_action=enrich but already enriched (manual override path)
        can_draft = False
        if next_action == "draft":
            can_draft = True
        elif next_action == "enrich" and row.get("enrichment_notes"):
            # Already enriched but still marked enrich - allow manual draft
            can_draft = True

        if not can_draft:
            skipped += 1
            continue

        if status == "Drafted":
            skipped += 1
            continue

        # Merge enrichment_notes into effective headline for personalization
        # This allows draft to use enrichment data without changing fill_template signature
        enrichment_notes = row.get("enrichment_notes") or ""
        if enrichment_notes:
            # Create effective row with enrichment merged into headline for personalization
            effective_row = dict(row)
            current_headline = row.get("headline") or ""
            if current_headline:
                effective_row["headline"] = f"{current_headline} | {enrichment_notes}"
            else:
                effective_row["headline"] = enrichment_notes
        else:
            effective_row = row

        subject, body = fill_template(
            template_subject, template_body, effective_row, config
        )

        update_row(
            wb,
            row_id,
            {
                "draft_subject": subject,
                "draft_body": body,
                "status": "Drafted",
                "next_action": "review",
            },
        )
        drafted += 1

    save_workbook(wb, workbook_path)

    print(f"Draft complete: {drafted} drafted, {skipped} skipped")
    return {"drafted": drafted, "skipped": skipped}


def cmd_approve(workbook_path: str):
    """Auto-approve all drafted rows for testing.

    Sets status=Approved, next_action=send for all rows
    where status=Drafted and next_action=review.
    """
    wb = load_workbook(workbook_path)
    rows = read_all_rows(wb)

    approved = 0
    skipped = 0

    for row in rows:
        status = row.get("status") or ""
        next_action = row.get("next_action") or ""
        row_id = row.get("row_id")

        if status == "Drafted" and next_action == "review":
            update_row(
                wb,
                row_id,
                {
                    "status": "Approved",
                    "next_action": "send",
                },
            )
            approved += 1
        else:
            skipped += 1

    save_workbook(wb, workbook_path)

    print(f"Approve complete: {approved} approved, {skipped} skipped")
    return {"approved": approved, "skipped": skipped}


def cmd_summary(workbook_path: str):
    """Print summary counts by status and next_action."""
    wb = load_workbook(workbook_path)
    rows = read_all_rows(wb)

    status_counts = Counter(row.get("status") or "(empty)" for row in rows)
    action_counts = Counter(row.get("next_action") or "(empty)" for row in rows)

    print("\n=== Status Counts ===")
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count}")

    print("\n=== Next Action Counts ===")
    for action, count in sorted(action_counts.items()):
        print(f"  {action}: {count}")

    print(f"\nTotal rows: {len(rows)}")

    return {
        "status_counts": dict(status_counts),
        "action_counts": dict(action_counts),
        "total": len(rows),
    }


def usage():
    print(__doc__.strip(), file=sys.stderr)
    sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        usage()

    cmd = sys.argv[1]
    workbook_path = sys.argv[2]

    if cmd == "filter":
        if len(sys.argv) < 4:
            usage()
        config_path = sys.argv[3]
        # Check for --no-enrichment flag for backward compatibility
        use_enrichment = "--no-enrichment" not in sys.argv
        cmd_filter(workbook_path, config_path, use_enrichment=use_enrichment)
    elif cmd == "draft":
        if len(sys.argv) < 5:
            usage()
        config_path = sys.argv[3]
        template_path = sys.argv[4]
        cmd_draft(workbook_path, config_path, template_path)
    elif cmd == "approve":
        cmd_approve(workbook_path)
    elif cmd == "summary":
        cmd_summary(workbook_path)
    else:
        print(f"Unknown command: {cmd}", file=sys.stderr)
        usage()
