#!/usr/bin/env python3
"""Excel read/write utilities for linkedin-sourcing skill.

Usage:
    python3 excel_utils.py create   <path>
    python3 excel_utils.py read     <path> [--filter next_action=send]
    python3 excel_utils.py update   <path> <row_id> '{"status":"Sent"}'
    python3 excel_utils.py append   <path> '{"name":"John","company":"Google"}'
    python3 excel_utils.py count    <path> [--filter status=Sent,date_sent=2026-04-09]
"""

from __future__ import annotations

import sys
import json
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
SKILL_DIR = SCRIPT_DIR.parent

COLUMNS = [
    "row_id",
    "name",
    "company",
    "title",
    "profile_url",
    "est_yoe",
    "highest_degree",
    "school",
    "status",
    "next_action",
    "draft_subject",
    "draft_body",
    "date_sent",
    "attempts",
    "last_contact",
    "reply_type",
    "reply_summary",
    "notes",
    "headline",
    "location",
    "enrichment_notes",
    "enriched_at",
]

_openpyxl = None


def get_sheet_headers(ws):
    """Return the non-empty header values from the first sheet row."""
    return [cell.value for cell in ws[1] if cell.value is not None]


def get_openpyxl():
    """Lazy-load and cache openpyxl, auto-installing if missing."""
    global _openpyxl
    if _openpyxl is None:
        try:
            import openpyxl
        except ImportError:
            import subprocess

            # Try normal install first, fall back to --break-system-packages
            # for macOS managed Python (PEP 668)
            try:
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
                )
            except subprocess.CalledProcessError:
                subprocess.check_call(
                    [
                        sys.executable,
                        "-m",
                        "pip",
                        "install",
                        "openpyxl",
                        "-q",
                        "--break-system-packages",
                    ],
                )
            import openpyxl
        _openpyxl = openpyxl
    return _openpyxl


def create(path):
    """Create a new Excel workbook with the Candidates sheet and header row."""
    wb = get_openpyxl().Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.append(COLUMNS)
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"Created {path}")


def load_candidates_workbook(path, migrate_schema=True):
    """Load the Candidates workbook, optionally applying schema migration first."""
    if migrate_schema and Path(path).exists():
        ensure_schema(path)
    return get_openpyxl().load_workbook(path)


def read(path, filters=None):
    """Read all candidate rows, optionally filtered by column=value pairs.

    Handles both old (18-column) and new (20-column) schemas gracefully.
    Missing columns return None instead of raising errors.
    """
    wb = load_candidates_workbook(path)
    ws = wb["Candidates"]

    # Read actual headers from the workbook (handles old/new schemas)
    actual_headers = get_sheet_headers(ws)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[0].value is None:
            continue
        # Map by actual header position, default to None for missing columns
        data = {}
        for i, header in enumerate(actual_headers):
            data[header] = row[i].value if i < len(row) else None
        # Ensure all canonical columns exist (as None if missing)
        for col in COLUMNS:
            if col not in data:
                data[col] = None
        if filters and not all(str(data.get(k, "")) == v for k, v in filters.items()):
            continue
        rows.append(data)
    return rows


def update(path, row_id, updates):
    """Update a single row by row_id with the given column values."""
    wb = load_candidates_workbook(path)
    ws = wb["Candidates"]
    actual_headers = get_sheet_headers(ws)
    header_to_index = {header: idx for idx, header in enumerate(actual_headers)}
    row_id = int(row_id)
    found = False
    for row in ws.iter_rows(min_row=2):
        if row[0].value == row_id:
            for col_name, value in updates.items():
                col_idx = header_to_index.get(col_name)
                if col_idx is not None:
                    row[col_idx].value = value
            found = True
            break
    if not found:
        print(f"Row {row_id} not found", file=sys.stderr)
        sys.exit(1)
    wb.save(path)
    print(f"Updated row {row_id}")


def append(path, data):
    """Append a new candidate row and return the assigned row_id."""
    wb = load_candidates_workbook(path)
    ws = wb["Candidates"]
    actual_headers = get_sheet_headers(ws)
    next_id = ws.max_row  # row 1 is header, so max_row gives next sequential id
    row_data = dict(data)
    row_data["row_id"] = next_id
    ws.append([row_data.get(col) for col in actual_headers])
    wb.save(path)
    return next_id


def upsert(path, data, key_column="profile_url"):
    """Upsert a candidate row based on a key column.

    If a row with the same key_column value exists, update it.
    Otherwise, append a new row.

    Args:
        path: Path to the workbook
        data: Dict of column values to upsert
        key_column: Column to use as unique key (default: profile_url)

    Returns:
        Dict with:
            - row_id: The row ID (existing or new)
            - action: "updated" or "inserted"
    """
    wb = load_candidates_workbook(path)
    ws = wb["Candidates"]
    actual_headers = get_sheet_headers(ws)
    header_to_index = {header: idx for idx, header in enumerate(actual_headers)}

    key_value = data.get(key_column)
    key_col_idx = header_to_index.get(key_column)

    # Search for existing row by key
    if key_value and key_col_idx is not None:
        for row in ws.iter_rows(min_row=2):
            if row[key_col_idx].value == key_value:
                # Found existing row - update it
                row_id = row[0].value  # row_id is first column
                for col_name, value in data.items():
                    col_idx = header_to_index.get(col_name)
                    if col_idx is not None:
                        row[col_idx].value = value
                wb.save(path)
                return {"row_id": row_id, "action": "updated"}

    # No existing row found - append new
    next_id = ws.max_row
    row_data = dict(data)
    row_data["row_id"] = next_id
    ws.append([row_data.get(col) for col in actual_headers])
    wb.save(path)
    return {"row_id": next_id, "action": "inserted"}


def get_existing_keys(path, key_column="profile_url"):
    """Get a set of all values for a key column (for deduplication).

    Args:
        path: Path to the workbook
        key_column: Column to extract values from (default: profile_url)

    Returns:
        Set of non-empty values for the key column
    """
    rows = read(path)
    keys = set()
    for row in rows:
        value = row.get(key_column)
        if value:
            keys.add(value)
    return keys


def count(path, filters=None):
    """Count rows matching the given filters."""
    return len(read(path, filters))


def ensure_schema(path):
    """Ensure workbook has all columns defined in COLUMNS.

    Safely migrates old workbooks by adding missing columns without data loss.
    Returns True if migration was performed, False if no changes needed.
    """
    wb = get_openpyxl().load_workbook(path)
    ws = wb["Candidates"]

    # Get current headers from first row
    current_headers = []
    for cell in ws[1]:
        if cell.value is not None:
            current_headers.append(cell.value)

    # Check if migration is needed
    missing_cols = [col for col in COLUMNS if col not in current_headers]
    if not missing_cols:
        return False

    # Add missing columns to header row
    for col_name in missing_cols:
        next_col_idx = len(current_headers) + 1
        ws.cell(row=1, column=next_col_idx, value=col_name)
        current_headers.append(col_name)

    wb.save(path)
    return True


def get_column_index(col_name):
    """Get the 0-based index of a column by name."""
    if col_name in COLUMNS:
        return COLUMNS.index(col_name)
    return None


def parse_filters(args):
    """Parse --filter key=val,key=val from CLI args."""
    filters = {}
    for arg in args:
        if arg == "--filter":
            continue
        if "=" in arg:
            for pair in arg.split(","):
                k, v = pair.split("=", 1)
                filters[k.strip()] = v.strip()
    return filters or None


def usage():
    print(__doc__.strip(), file=sys.stderr)
    sys.exit(1)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        usage()

    cmd, path = sys.argv[1], sys.argv[2]

    if cmd == "create":
        create(path)
    elif cmd == "read":
        print(json.dumps(read(path, parse_filters(sys.argv[3:])), default=str))
    elif cmd == "update":
        if len(sys.argv) < 5:
            usage()
        update(path, sys.argv[3], json.loads(sys.argv[4]))
    elif cmd == "append":
        if len(sys.argv) < 4:
            usage()
        print(append(path, json.loads(sys.argv[3])))
    elif cmd == "count":
        print(count(path, parse_filters(sys.argv[3:])))
    else:
        print(f"Unknown command: {cmd}", file=sys.stderr)
        usage()
